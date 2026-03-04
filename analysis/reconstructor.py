"""
资产负债表重构模块

将资产负债表的各个项目按照财务分析需求重新分类归集，
生成资产资本表（Asset-Capital Statement），便于进行深度财务分析。

输出格式：
- 详细列出所有项目（而非合并指标）
- 只包含年度数据（12-31）+ TTM数据
- 横向逆序排列（最新的在左边）
"""
import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, List
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine


class BalanceSheetReconstructor:
    """资产负债表重构器"""
    
    def __init__(self, db_path: str = 'database.sqlite'):
        """
        初始化重构器
        
        Args:
            db_path: 数据库文件路径
        """
        self.engine = create_engine(f'sqlite:///{db_path}')
        
    def load_balance_sheet_data(self, stock_code: str, years: int = 10) -> pd.DataFrame:
        """
        从数据库加载资产负债表数据，只加载年度数据（12-31）
        
        Args:
            stock_code: 股票代码（不带前缀，如'600519'）
            years: 加载最近N年的数据
            
        Returns:
            资产负债表DataFrame，按report_date排序
        """
        query = f"""
        SELECT * FROM balance_sheets 
        WHERE stock_code = '{stock_code}'
        AND substr(report_date, 6, 5) = '12-31'
        ORDER BY report_date DESC
        LIMIT {years}
        """
        df = pd.read_sql(query, self.engine)
        df['report_date'] = pd.to_datetime(df['report_date'])
        return df.sort_values('report_date', ascending=False)
    
    def calculate_financial_assets(self, row: pd.Series) -> float:
        """
        计算金融资产合计
        
        规则1: 金融资产合计 = 货币资金 + 拆出资金 + 交易性金融资产 + 衍生金融资产 + 
               买入返售金融资产 + 持有待售资产 + 一年内到期的非流动资产 + 
               发放贷款及垫款 + 债权投资 + 其他债权投资 + 其他权益工具投资 + 
               其他非流动金融资产 + 投资性房地产
        """
        fields = [
            'monetary_capital',  # 货币资金
            'loans_to_other_banks',  # 拆出资金
            'trading_financial_assets',  # 交易性金融资产
            'derivative_financial_assets',  # 衍生金融资产
            'financial_assets_purchased_for_resale',  # 买入返售金融资产
            'assets_held_for_sale',  # 持有待售资产
            'non_current_assets_due_within_one_year',  # 一年内到期的非流动资产
            'loans_and_advances',  # 发放贷款及垫款
            'debt_investments',  # 债权投资
            'other_debt_investments',  # 其他债权投资
            'other_equity_instrument_investments',  # 其他权益工具投资
            'other_non_current_financial_assets',  # 其他非流动金融资产
            'investment_property',  # 投资性房地产
        ]
        total = 0.0
        for f in fields:
            value = row.get(f)
            if pd.notna(value):
                total += float(value)
        return total
    
    def calculate_long_term_equity_investment(self, row: pd.Series) -> float:
        """
        计算长期股权投资
        
        规则2: 长期股权投资（单独一项）
        """
        value = row.get('long_term_equity_investments')
        return float(value) if pd.notna(value) else 0.0
    
    def calculate_operating_assets(self, row: pd.Series) -> float:
        """
        计算营运资产小计
        
        规则3: 营运资产小计 = 应收票据 + 应收账款 + 应收款项融资 + 预付款项 + 
               其他应收款 + 存货 + 合同资产 + 长期应收款 + 其他流动资产 + 
               结算备付金 + 应收保费 + 应收分保账款 + 应收分保合同准备金 + 保险合同准备金
        """
        fields = [
            'notes_receivable',  # 应收票据
            'accounts_receivable',  # 应收账款
            'receivables_financing',  # 应收款项融资
            'prepayments',  # 预付款项
            'other_receivables',  # 其他应收款
            'inventories',  # 存货
            'contract_assets',  # 合同资产
            'long_term_receivables',  # 长期应收款
            'other_current_assets',  # 其他流动资产
            'settlement_provisions',  # 结算备付金
            'insurance_premiums_receivable',  # 应收保费
            'reinsurance_receivables',  # 应收分保账款
            'reinsurance_contract_reserves_receivable',  # 应收分保合同准备金
            'insurance_contract_reserves',  # 保险合同准备金
        ]
        total = 0.0
        for f in fields:
            value = row.get(f)
            if pd.notna(value):
                total += float(value)
        return total
    
    def calculate_operating_liabilities(self, row: pd.Series) -> float:
        """
        计算营运负债小计
        
        规则4: 营运负债小计 = 应付票据 + 应付账款 + 预收账款 + 合同负债 + 
               应付职工薪酬 + 长期应付职工薪酬 + 应交税费 + 其他应付款 + 
               应付手续费及佣金 + 应付分保账款 + 其他流动负债 + 预计流动负债 + 
               预计非流动负债 + 一年内到期的递延收益 + 长期递延收益 + 其他非流动负债
        """
        fields = [
            'notes_payable',  # 应付票据
            'accounts_payable',  # 应付账款
            'advances_from_customers',  # 预收账款
            'contract_liabilities',  # 合同负债
            'employee_benefits_payable',  # 应付职工薪酬
            'long_term_employee_benefits_payable',  # 长期应付职工薪酬
            'taxes_payable',  # 应交税费
            'other_payables',  # 其他应付款
            'fees_and_commissions_payable',  # 应付手续费及佣金
            'reinsurance_payables',  # 应付分保账款
            'other_current_liabilities',  # 其他流动负债
            'estimated_current_liabilities',  # 预计流动负债
            'estimated_non_current_liabilities',  # 预计非流动负债
            'deferred_revenue_due_within_one_year',  # 一年内到期的递延收益
            'long_term_deferred_revenue',  # 长期递延收益
            'other_non_current_liabilities',  # 其他非流动负债
        ]
        total = 0.0
        for f in fields:
            value = row.get(f)
            if pd.notna(value):
                total += float(value)
        return total
    
    def calculate_cyclical_operating_investment(self, operating_assets: float, operating_liabilities: float) -> float:
        """
        计算周期性经营投入合计
        
        规则5: 周期性经营投入合计 = 营运资产小计 - 营运负债小计
        """
        return operating_assets - operating_liabilities
    
    def calculate_long_term_operating_assets(self, row: pd.Series) -> float:
        """
        计算长期经营资产合计
        
        规则6: 长期经营资产合计 = 固定资产 + 在建工程 + 生产性生物资产 + 
               消耗性生物资产 + 油气资产 + 使用权资产 + 无形资产 + 开发支出 + 
               商誉 + 长期待摊费用 + 其他非流动资产 + 递延所得税资产 - 递延所得税负债
        """
        fields_positive = [
            'fixed_assets_net',  # 固定资产
            'construction_in_progress',  # 在建工程
            'productive_biological_assets',  # 生产性生物资产
            'consumptive_biological_assets',  # 消耗性生物资产
            'oil_and_gas_assets',  # 油气资产
            'right_of_use_assets',  # 使用权资产
            'intangible_assets',  # 无形资产
            'development_expenditure',  # 开发支出
            'goodwill',  # 商誉
            'long_term_deferred_expenses',  # 长期待摊费用
            'other_non_current_assets',  # 其他非流动资产
            'deferred_tax_assets',  # 递延所得税资产
        ]
        fields_negative = [
            'deferred_tax_liabilities',  # 递延所得税负债
        ]
        
        total = 0.0
        for f in fields_positive:
            value = row.get(f)
            if pd.notna(value):
                total += float(value)
        for f in fields_negative:
            value = row.get(f)
            if pd.notna(value):
                total -= float(value)
        return total
    
    def calculate_total_operating_assets(self, cyclical: float, long_term: float) -> float:
        """
        计算经营资产合计
        
        规则7: 经营资产合计 = 周期性经营投入合计 + 长期经营资产合计
        """
        return cyclical + long_term
    
    def calculate_short_term_debt(self, row: pd.Series) -> float:
        """
        计算短期债务
        
        规则8: 短期债务 = 短期借款 + 向中央银行借款 + 拆入资金 + 交易性金融负债 + 
               衍生金融负债 + 卖出回购金融资产 + 吸收存款及同业存放 + 
               代理买卖证券款 + 代理承销证券款 + 应付利息 + 持有待售负债 + 
               一年内到期的非流动负债
        """
        fields = [
            'short_term_borrowings',  # 短期借款
            'borrowings_from_central_bank',  # 向中央银行借款
            'borrowings_from_other_banks',  # 拆入资金
            'trading_financial_liabilities',  # 交易性金融负债
            'derivative_financial_liabilities',  # 衍生金融负债
            'financial_assets_sold_for_repurchase',  # 卖出回购金融资产
            'deposits_from_customers_and_banks',  # 吸收存款及同业存放
            'securities_trading_agency_payables',  # 代理买卖证券款
            'securities_underwriting_agency_payables',  # 代理承销证券款
            'interest_payable',  # 应付利息
            'liabilities_held_for_sale',  # 持有待售负债
            'non_current_liabilities_due_within_one_year',  # 一年内到期的非流动负债
        ]
        total = 0.0
        for f in fields:
            value = row.get(f)
            if pd.notna(value):
                total += float(value)
        return total
    
    def calculate_long_term_debt(self, row: pd.Series) -> float:
        """
        计算长期债务
        
        规则9: 长期债务 = 长期借款 + 应付债券 + 租赁负债 + 长期应付款
        """
        fields = [
            'long_term_borrowings',  # 长期借款
            'bonds_payable',  # 应付债券
            'lease_liabilities',  # 租赁负债
            'long_term_payables',  # 长期应付款
        ]
        total = 0.0
        for f in fields:
            value = row.get(f)
            if pd.notna(value):
                total += float(value)
        return total
    
    def calculate_total_interest_bearing_debt(self, short_term: float, long_term: float) -> float:
        """
        计算有息债务合计
        
        规则10: 有息债务合计 = 短期债务 + 长期债务
        """
        return short_term + long_term
    
    def calculate_parent_equity(self, row: pd.Series) -> float:
        """
        计算归属于母公司股东权益合计
        
        规则11: 优先使用数据库中已计算的 equity_attributable_to_parent_company 字段
                如果该字段为空，则使用 total_owners_equity - minority_interests
        """
        # 优先使用数据库中已计算好的归属于母公司股东权益
        parent_equity = row.get('equity_attributable_to_parent_company')
        
        if pd.notna(parent_equity) and parent_equity != 0:
            return parent_equity
        
        # 如果没有，则用总权益减去少数股东权益
        total_equity = row.get('total_owners_equity', 0) or 0
        minority = row.get('minority_interests', 0) or 0
        return total_equity - minority
    
    def calculate_total_equity(self, row: pd.Series) -> float:
        """
        计算所有者权益合计
        
        规则12: 直接使用数据库中的 total_owners_equity 字段
        """
        value = row.get('total_owners_equity')
        return float(value) if pd.notna(value) else 0.0
    
    def get_structured_items(self) -> List[tuple]:
        """
        获取结构化的项目列表，按照附图的顺序和分组
        返回格式：[(显示名称, 数据库字段名或计算类型, 是否为标题/小计), ...]
        """
        return [
            # 资产类标题
            ('资产类', 'SECTION_TITLE', True),
            (' ', 'BLANK', True),
            
            # 金融资产
            ('金融资产', 'SECTION_TITLE', True),
            ('货币资金', 'monetary_capital', False),
            ('拆出资金', 'loans_to_other_banks', False),
            ('交易性金融资产', 'trading_financial_assets', False),
            ('衍生金融资产', 'derivative_financial_assets', False),
            ('买入返售金融资产', 'financial_assets_purchased_for_resale', False),
            ('持有待售资产', 'assets_held_for_sale', False),
            ('一年内到期的非流动资产', 'non_current_assets_due_within_one_year', False),
            ('发放贷款及垫款', 'loans_and_advances', False),
            ('债权投资', 'debt_investments', False),
            ('其他债权投资', 'other_debt_investments', False),
            ('其他权益工具投资', 'other_equity_instrument_investments', False),
            ('其他非流动金融资产', 'other_non_current_financial_assets', False),
            ('投资性房地产', 'investment_property', False),
            ('金融资产合计', 'CALC_FINANCIAL_ASSETS', True),
            (' ', 'BLANK', True),
            
            # 长期股权投资
            ('长期股权投资', 'long_term_equity_investments', False),
            (' ', 'BLANK', True),
            
            # 营运资产
            ('营运资产', 'SECTION_TITLE', True),
            ('应收票据', 'notes_receivable', False),
            ('应收账款', 'accounts_receivable', False),
            ('应收款项融资', 'receivables_financing', False),
            ('预付款项', 'prepayments', False),
            ('其他应收款', 'other_receivables', False),
            ('存货', 'inventories', False),
            ('合同资产', 'contract_assets', False),
            ('长期应收款', 'long_term_receivables', False),
            ('其他流动资产', 'other_current_assets', False),
            ('营运资产小计', 'CALC_OPERATING_ASSETS', True),
            (' ', 'BLANK', True),
            
            # 营运负债
            ('营运负债', 'SECTION_TITLE', True),
            ('应付票据', 'notes_payable', False),
            ('应付账款', 'accounts_payable', False),
            ('预收账款', 'advances_from_customers', False),
            ('合同负债', 'contract_liabilities', False),
            ('应付职工薪酬', 'employee_benefits_payable', False),
            ('应交税费', 'taxes_payable', False),
            ('其他应付款', 'other_payables', False),
            ('其他流动负债', 'other_current_liabilities', False),
            ('营运负债小计', 'CALC_OPERATING_LIABILITIES', True),
            (' ', 'BLANK', True),
            
            # 周期性经营投入
            ('周期性经营投入合计', 'CALC_CYCLICAL_INVESTMENT', True),
            (' ', 'BLANK', True),
            
            # 长期经营资产
            ('长期经营资产', 'SECTION_TITLE', True),
            ('固定资产', 'fixed_assets_net', False),
            ('在建工程', 'construction_in_progress', False),
            ('生产性生物资产', 'productive_biological_assets', False),
            ('消耗性生物资产', 'consumptive_biological_assets', False),
            ('油气资产', 'oil_and_gas_assets', False),
            ('使用权资产', 'right_of_use_assets', False),
            ('无形资产', 'intangible_assets', False),
            ('开发支出', 'development_expenditure', False),
            ('商誉', 'goodwill', False),
            ('长期待摊费用', 'long_term_deferred_expenses', False),
            ('其他非流动资产', 'other_non_current_assets', False),
            ('递延所得税资产', 'deferred_tax_assets', False),
            ('减：递延所得税负债', 'deferred_tax_liabilities', False),
            ('长期经营资产合计', 'CALC_LONG_TERM_OPERATING', True),
            (' ', 'BLANK', True),
            
            # 经营资产合计
            ('经营资产合计', 'CALC_TOTAL_OPERATING', True),
            (' ', 'BLANK', True),
            
            # 资产总额
            ('资产总额', 'CALC_TOTAL_ASSETS', True),
            (' ', 'BLANK', True),
            
            # 短期债务
            ('短期债务', 'SECTION_TITLE', True),
            ('短期借款', 'short_term_borrowings', False),
            ('向中央银行借款', 'borrowings_from_central_bank', False),
            ('拆入资金', 'borrowings_from_other_banks', False),
            ('交易性金融负债', 'trading_financial_liabilities', False),
            ('一年内到期的非流动负债', 'non_current_liabilities_due_within_one_year', False),
            ('短期债务合计', 'CALC_SHORT_TERM_DEBT', True),
            (' ', 'BLANK', True),
            
            # 长期债务
            ('长期债务', 'SECTION_TITLE', True),
            ('长期借款', 'long_term_borrowings', False),
            ('应付债券', 'bonds_payable', False),
            ('租赁负债', 'lease_liabilities', False),
            ('长期应付款', 'long_term_payables', False),
            ('长期债务合计', 'CALC_LONG_TERM_DEBT', True),
            (' ', 'BLANK', True),
            
            # 有息债务合计
            ('有息债务合计', 'CALC_TOTAL_DEBT', True),
            (' ', 'BLANK', True),
            
            # 权益
            ('权益', 'SECTION_TITLE', True),
            ('股本', 'paid_in_capital', False),
            ('资本公积', 'capital_reserve', False),
            ('减：库存股', 'less_treasury_stock', False),
            ('其他综合收益', 'other_comprehensive_income', False),
            ('盈余公积', 'surplus_reserve', False),
            ('一般风险准备', 'general_risk_reserve', False),
            ('未分配利润', 'retained_earnings', False),
            ('归属于母公司股东权益合计', 'CALC_PARENT_EQUITY', True),
            ('少数股东权益', 'minority_interests', False),
            ('所有者权益合计', 'CALC_TOTAL_EQUITY', True),
        ]
    
    def load_latest_quarter_data(self, stock_code: str) -> Optional[pd.Series]:
        """
        加载最新一个季度的数据（用于计算TTM）
        """
        query = f"""
        SELECT * FROM balance_sheets 
        WHERE stock_code = '{stock_code}'
        ORDER BY report_date DESC
        LIMIT 1
        """
        df = pd.read_sql(query, self.engine)
        if df.empty:
            return None
        df['report_date'] = pd.to_datetime(df['report_date'])
        return df.iloc[0]
    
    def reconstruct_balance_sheet(self, stock_code: str, years: int = 10) -> pd.DataFrame:
        """
        重构资产负债表，输出详细项目列表（包含分组标题和小计）
        
        Args:
            stock_code: 股票代码
            years: 分析年限
            
        Returns:
            重构后的资产资本表DataFrame，包含年度数据和TTM数据
        """
        # 加载年度数据（12-31）
        df_annual = self.load_balance_sheet_data(stock_code, years)
        
        if df_annual.empty:
            return pd.DataFrame()
        
        # 加载最新季度数据（用于TTM）
        latest_quarter = self.load_latest_quarter_data(stock_code)
        
        # 获取结构化项目列表
        structured_items = self.get_structured_items()
        
        # 构建结果字典：{期间: {项目名: 值}}
        results = {}
        
        # 处理年度数据列（逆序）
        for _, row in df_annual.iterrows():
            year = row['report_date'].year
            col_name = f"{year}Q4"
            results[col_name] = self._calculate_all_items(row, structured_items)
        
        # 添加TTM列（如果最新季度不是12-31）
        if latest_quarter is not None:
            latest_date = latest_quarter['report_date']
            if latest_date.month != 12:
                year = latest_date.year
                quarter = (latest_date.month - 1) // 3 + 1
                col_name = f"{year}Q{quarter}-TTM"
                ttm_data = self._calculate_all_items(latest_quarter, structured_items)
                # TTM列插入到最前面
                results = {col_name: ttm_data, **results}
        
        # 转换为DataFrame
        df_result = pd.DataFrame(results)
        
        return df_result
    
    def _calculate_all_items(self, row: pd.Series, structured_items: List[tuple]) -> Dict[str, float]:
        """
        计算一个报告期的所有项目值（包括小计）
        
        Args:
            row: 数据库中的一行数据
            structured_items: 结构化项目列表
            
        Returns:
            {项目名: 值} 的字典
        """
        result = {}
        
        # 用于存储中间计算结果
        temp_values = {}
        
        for display_name, field_or_calc, is_summary in structured_items:
            if field_or_calc == 'BLANK':
                result[display_name] = None
            elif field_or_calc == 'SECTION_TITLE':
                result[display_name] = None
            elif field_or_calc.startswith('CALC_'):
                # 计算小计
                calc_value = self._calculate_subtotal(field_or_calc, row, temp_values)
                result[display_name] = calc_value
                temp_values[field_or_calc] = calc_value
            else:
                # 从数据库读取字段值
                value = row.get(field_or_calc)
                result[display_name] = float(value) if pd.notna(value) else 0.0
        
        return result
    
    def _calculate_subtotal(self, calc_type: str, row: pd.Series, temp_values: Dict[str, float]) -> float:
        """
        计算小计值
        
        Args:
            calc_type: 计算类型（如CALC_FINANCIAL_ASSETS）
            row: 数据库中的一行数据
            temp_values: 已计算的中间值
            
        Returns:
            计算结果
        """
        if calc_type == 'CALC_FINANCIAL_ASSETS':
            return self.calculate_financial_assets(row)
        elif calc_type == 'CALC_OPERATING_ASSETS':
            return self.calculate_operating_assets(row)
        elif calc_type == 'CALC_OPERATING_LIABILITIES':
            return self.calculate_operating_liabilities(row)
        elif calc_type == 'CALC_CYCLICAL_INVESTMENT':
            operating_assets = temp_values.get('CALC_OPERATING_ASSETS', 0.0)
            operating_liabilities = temp_values.get('CALC_OPERATING_LIABILITIES', 0.0)
            return operating_assets - operating_liabilities
        elif calc_type == 'CALC_LONG_TERM_OPERATING':
            return self.calculate_long_term_operating_assets(row)
        elif calc_type == 'CALC_TOTAL_OPERATING':
            cyclical = temp_values.get('CALC_CYCLICAL_INVESTMENT', 0.0)
            long_term = temp_values.get('CALC_LONG_TERM_OPERATING', 0.0)
            return cyclical + long_term
        elif calc_type == 'CALC_TOTAL_ASSETS':
            # 资产总额 = 金融资产合计 + 长期股权投资 + 经营资产合计
            financial_assets = temp_values.get('CALC_FINANCIAL_ASSETS', 0.0)
            long_term_equity = row.get('long_term_equity_investments')
            long_term_equity = float(long_term_equity) if pd.notna(long_term_equity) else 0.0
            operating_assets = temp_values.get('CALC_TOTAL_OPERATING', 0.0)
            return financial_assets + long_term_equity + operating_assets
        elif calc_type == 'CALC_SHORT_TERM_DEBT':
            return self.calculate_short_term_debt(row)
        elif calc_type == 'CALC_LONG_TERM_DEBT':
            return self.calculate_long_term_debt(row)
        elif calc_type == 'CALC_TOTAL_DEBT':
            short_term = temp_values.get('CALC_SHORT_TERM_DEBT', 0.0)
            long_term = temp_values.get('CALC_LONG_TERM_DEBT', 0.0)
            return short_term + long_term
        elif calc_type == 'CALC_PARENT_EQUITY':
            return self.calculate_parent_equity(row)
        elif calc_type == 'CALC_TOTAL_EQUITY':
            return self.calculate_total_equity(row)
        else:
            return 0.0
    
    def get_section_titles(self) -> List[tuple]:
        """
        获取分组标题列表
        返回格式：[(标题名称, 起始行号), ...]
        """
        return [
            ('资产类', 1),
            ('金融资产', 2),
            ('长期股权投资', 15),
            ('营运资产', 16),
            ('营运负债', 25),
            ('长期经营资产', 33),
            ('短期债务', 46),
            ('长期债务', 51),
            ('权益', 55),
        ]
    
    def export_to_excel(
        self,
        reconstructed_df: pd.DataFrame,
        stock_code: str,
        output_dir: str = 'output'
    ) -> str:
        """
        导出资产资本表到Excel文件
        
        Args:
            reconstructed_df: 重构后的数据（已经是行=项目，列=期间的格式）
            stock_code: 股票代码
            output_dir: 输出目录
            
        Returns:
            生成的Excel文件路径
        """
        if reconstructed_df.empty:
            raise ValueError("数据为空，无法导出")
        
        # 创建输出目录
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        
        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{stock_code}_资产资本表_{timestamp}.xlsx"
        filepath = Path(output_dir) / filename
        
        # 写入Excel
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            reconstructed_df.to_excel(writer, sheet_name='资产资本表')
            
            # 获取工作表
            workbook = writer.book
            worksheet = writer.sheets['资产资本表']
            
            # 应用格式
            self._apply_excel_formatting(worksheet, reconstructed_df)
        
        return str(filepath)
    
    def _format_quarter(self, date: pd.Timestamp) -> str:
        """
        将日期格式化为季度格式（如2019Q1）
        
        Args:
            date: 日期
            
        Returns:
            格式化后的季度字符串
        """
        year = date.year
        month = date.month
        
        if month in [1, 2, 3]:
            quarter = 'Q1'
        elif month in [4, 5, 6]:
            quarter = 'Q2'
        elif month in [7, 8, 9]:
            quarter = 'Q3'
        else:
            quarter = 'Q4'
        
        return f"{year}{quarter}"
    
    def _apply_excel_formatting(self, worksheet, df: pd.DataFrame):
        """
        应用Excel格式化（参考资产资本表参考.xlsx的格式）
        
        Args:
            worksheet: openpyxl工作表对象
            df: 数据DataFrame
        """
        from openpyxl.styles.colors import Color
        
        # 定义样式 - 参考Excel使用的主题颜色
        # 表头：深色背景（Theme 0, Tint -0.15）
        header_font = Font(name='等线', size=11, bold=True)
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # 大标题（如"资产类"）：浅蓝色背景（Theme 5, Tint 0.4）
        section_title_font = Font(name='等线', size=11, bold=True)
        section_title_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')  # 橙色
        section_title_alignment = Alignment(horizontal='left', vertical='center')
        
        # 子标题（如"金融资产"）：更浅的蓝色（Theme 5, Tint 0.8）
        subsection_font = Font(name='等线', size=11, bold=True)
        subsection_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # 浅橙色
        subsection_alignment = Alignment(horizontal='left', vertical='center')
        
        # 小计行：与子标题相同的浅蓝色
        subtotal_font = Font(name='等线', size=11, bold=True)
        subtotal_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # 黄色
        subtotal_alignment = Alignment(horizontal='left', vertical='center')
        
        # 普通项目：无背景
        normal_font = Font(name='等线', size=11, bold=False)
        normal_alignment = Alignment(horizontal='left', vertical='center')
        
        # 数据单元格
        data_font = Font(name='等线', size=11)
        data_alignment = Alignment(horizontal='right', vertical='center')
        
        # 边框
        border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        
        # 格式化表头（第一行）
        for col_num in range(1, len(df.columns) + 2):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 格式化索引列（第一列）- 根据行类型应用不同格式
        for row_num in range(2, len(df) + 2):
            cell = worksheet.cell(row=row_num, column=1)
            row_name = str(df.index[row_num - 2])
            
            # 判断行类型并应用相应格式
            if row_name in ['资产类', '资本类']:
                # 大标题
                cell.font = section_title_font
                cell.fill = section_title_fill
                cell.alignment = section_title_alignment
            elif row_name in ['金融资产', '营运资产', '营运负债', '长期经营资产', '短期债务', '长期债务', '权益']:
                # 子标题
                cell.font = subsection_font
                cell.fill = subsection_fill
                cell.alignment = subsection_alignment
            elif '合计' in row_name or '小计' in row_name or row_name == '资产总额':
                # 小计行
                cell.font = subtotal_font
                cell.fill = subtotal_fill
                cell.alignment = subtotal_alignment
            elif row_name.strip() == '':
                # 空行 - 无格式
                continue
            else:
                # 普通项目
                cell.font = normal_font
                cell.alignment = normal_alignment
            
            cell.border = border
        
        # 格式化数据单元格
        for row_num in range(2, len(df) + 2):
            for col_num in range(2, len(df.columns) + 2):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
                
                # 应用千位分隔符格式（参考Excel的格式）
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = '_ * #,##0_ ;_ * -#,##0_ ;_ * "-"??_ ;_ @_ '
        
        # 调整列宽
        worksheet.column_dimensions['A'].width = 30  # 指标名称列
        for col_num in range(2, len(df.columns) + 2):
            col_letter = get_column_letter(col_num)
            worksheet.column_dimensions[col_letter].width = 18
        
        # 调整行高
        for row_num in range(1, len(df) + 2):
            worksheet.row_dimensions[row_num].height = 20
        
        # 冻结首行首列
        worksheet.freeze_panes = 'B2'
